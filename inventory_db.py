import sqlite3
import qrcode
import csv
import os
import shutil
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

class InventoryManagement:
    def __init__(self):
        """Initializes the InventoryManagement class and creates the database."""
        self.create_database()

    def create_database(self):
        """Creates the SQLite database and the products table if it doesn't exist."""
        conn = sqlite3.connect('inventory_database.db')
        cursor = conn.cursor()
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS products (
            id INTEGER PRIMARY KEY,
            product_name TEXT NOT NULL,
            quantity INTEGER NOT NULL,
            price REAL NOT NULL,
            description TEXT
        )
        ''')
        conn.commit()
        conn.close()

    def connect_db(self):
        """Establishes a connection to the SQLite database.

        Returns:
            tuple: A tuple containing the connection and cursor objects.
        """
        conn = sqlite3.connect('inventory_database.db')
        cursor = conn.cursor()
        return conn, cursor

    def add_product(self, product_name, quantity, price, description, product_id=None):
        """Adds a new product to the database.

        Args:
            product_name (str): Name of the product.
            quantity (int): Initial quantity of the product.
            price (float): Price of the product.
            description (str): Description of the product.
            product_id (int, optional): ID of the product; if provided, inserts with this ID.
        """
        conn, cursor = self.connect_db()
        if product_id:
            cursor.execute('''
            INSERT INTO products (id, product_name, quantity, price, description)
            VALUES (?, ?, ?, ?, ?)
            ''', (product_id, product_name, quantity, price, description))
        else:
            cursor.execute('''
            INSERT INTO products (product_name, quantity, price, description)
            VALUES (?, ?, ?, ?)
            ''', (product_name, quantity, price, description))
        conn.commit()
        conn.close()

    def add_stock(self, product_id, quantity):
        """Adds stock to an existing product in the inventory.

        Args:
            product_id (int): ID of the product.
            quantity (int): Quantity to be added.

        Raises:
            ValueError: If the quantity to add is negative.
        """
        if quantity < 0:
            raise ValueError("Quantity to add must be non-negative.")

        conn, cursor = self.connect_db()
        cursor.execute('''
            UPDATE products 
            SET quantity = quantity + ? 
            WHERE id = ? 
        ''', (quantity, product_id))
        conn.commit()
        conn.close()
    
    def remove_stock(self, product_id, quantity):
        """Removes stock from an existing product in the inventory.

        Args:
            product_id (int): ID of the product.
            quantity (int): Quantity to be removed.

        Raises:
            ValueError: If the quantity to remove is negative, exceeds available stock, or if the product ID is not found.
        """
        if quantity < 0:
            raise ValueError("Quantity to remove must be non-negative.")

        conn, cursor = self.connect_db()
        cursor.execute('''
            SELECT quantity FROM products WHERE id = ? 
        ''', (product_id,))
        current_quantity = cursor.fetchone()

        if current_quantity is None:
            raise ValueError("Product ID not found.")

        current_quantity = current_quantity[0]

        if quantity > current_quantity:
            raise ValueError("Cannot remove more than available quantity.")
        
        cursor.execute('''
            UPDATE products 
            SET quantity = quantity - ? 
            WHERE id = ? 
        ''', (quantity, product_id))
        conn.commit()
        conn.close()

    def update_quantity(self, product_id, new_quantity):
        """Updates the quantity of a specific product.

        Args:
            product_id (int): ID of the product.
            new_quantity (int): New quantity for the product.
        """
        conn, cursor = self.connect_db()
        cursor.execute('''
        UPDATE products
        SET quantity = ?
        WHERE id = ?
        ''', (new_quantity, product_id))
        conn.commit()
        conn.close()

    def update_product(self, product_id, product_name, quantity, price, description):
        """Updates the details of an existing product.

        Args:
            product_id (int): ID of the product.
            product_name (str): New name of the product.
            quantity (int): New quantity of the product.
            price (float): New price of the product.
            description (str): New description of the product.
        """
        conn, cursor = self.connect_db()
        cursor.execute('''
        UPDATE products
        SET product_name = ?, quantity = ?, price = ?, description = ?
        WHERE id = ?
        ''', (product_name, quantity, price, description, product_id))
        conn.commit()
        conn.close()

    def delete_product(self, product_id):
        """Deletes a product from the inventory.

        Args:
            product_id (int): ID of the product to be deleted.
        """
        conn, cursor = self.connect_db()
        cursor.execute('''
        DELETE FROM products WHERE id = ?
        ''', (product_id,))
        conn.commit()
        conn.close()

    def get_product(self, product_id):
        """Retrieves the details of a specific product.

        Args:
            product_id (int): ID of the product.

        Returns:
            tuple: A tuple containing product details, or None if not found.
        """
        conn, cursor = self.connect_db()
        cursor.execute('''
        SELECT * FROM products WHERE id = ?
        ''', (product_id,))
        product = cursor.fetchone()
        conn.close()
        return product

    def search_products_by_name(self, name):
        """Searches for products by name.

        Args:
            name (str): Name or part of the name of the product to search for.

        Returns:
            list: A list of tuples containing products that match the search criteria.
        """
        conn, cursor = self.connect_db()
        cursor.execute('''
        SELECT * FROM products WHERE product_name LIKE ?
        ''', (f'%{name}%',))
        products = cursor.fetchall()
        conn.close()
        return products

    def list_products(self):
        """Lists all products in the inventory.

        Returns:
            list: A list of tuples containing all products in the inventory.
        """
        conn, cursor = self.connect_db()
        cursor.execute('''
        SELECT * FROM products
        ''')
        products = cursor.fetchall()
        conn.close()
        return products

    def generate_qr_code(self, product_data, product_id):
        """Generates a QR code for a specific product.

        Args:
            product_data (str): Data to be encoded in the QR code.
            product_id (int): ID of the product.
        """
        img = qrcode.make(product_data)
        img.save(f"./QRCodes/qr_code_{product_id}.png", "PNG")

    def process_outgoing_goods(self, filename):
        """Processes outgoing goods from a CSV file.

        Args:
            filename (str): Path to the CSV file containing outgoing goods data.
        """
        conn, cursor = self.connect_db()
        invoice_data = []

        with open(filename, newline='', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                product_data = row.get('CODECONTENT', None)
                product_quantity = int(row.get('QUANTITY', None))
                if product_data:
                    product_info = product_data.split(', ')
                    product_id = int(product_info[0])
                    product_name = product_info[1]
                    product_price = float(product_info[2])
                    cursor.execute('SELECT * FROM products WHERE id = ?', (product_id,))
                    existing_product = cursor.fetchone()

                    if existing_product:
                        new_quantity = existing_product[2] - product_quantity
                        new_quantity = max(new_quantity, 0)
                        cursor.execute('UPDATE products SET quantity = ? WHERE id = ?', (new_quantity, product_id))

                    # Add to invoice data
                    existing_invoice_item = next((item for item in invoice_data if item['Product ID'] == product_id), None)
                    if existing_invoice_item:
                        existing_invoice_item['Quantity'] += product_quantity
                        existing_invoice_item['Total Price'] += product_quantity * product_price
                    else:
                        invoice_data.append({
                            'Product ID': product_id,
                            'Product Name': product_name,
                            'Quantity': product_quantity,
                            'Unit Price': product_price,
                            'Total Price': product_quantity * product_price
                        })

        shutil.move(filename, "./OutgoingGoods/Processed")
        conn.commit()
        conn.close()
        self.write_invoice(invoice_data, filename)

    def write_invoice(self, invoice_data, filename):
        """Writes an invoice to an Excel file.

        Args:
            invoice_data (list): A list of dictionaries containing invoice data.
            filename (str): Original filename of the processed goods.
        """
        invoice_filename = f"Invoice_{os.path.basename(filename).replace('.csv', '')}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        invoice_path = os.path.join('./Invoices', invoice_filename)

        os.makedirs('./Invoices', exist_ok=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Invoice"

        # Formatting
        bold_font = Font(bold=True)
        center_alignment = Alignment(horizontal='center')
        border_style = Border(bottom=Side(style='thin'))

        # Header
        headers = ['Product ID', 'Product Name', 'Quantity', 'Unit Price', 'Total Price']
        ws.append(headers)
        for col in range(1, 6):
            cell = ws.cell(row=1, column=col)
            cell.font = bold_font
            cell.alignment = center_alignment
            cell.border = border_style

        # Data
        for data in invoice_data:
            ws.append([data['Product ID'], data['Product Name'], data['Quantity'], data['Unit Price'], data['Total Price']])

        # Adjust column width
        for col in ws.columns:
            max_length = max([len(str(cell.value)) for cell in col if cell.value])
            ws.column_dimensions[col[0].column_letter].width = max_length + 2

        # Format as currency
        for row in ws.iter_rows(min_row=2, max_col=5, max_row=len(invoice_data) + 1):
            row[3].number_format = '#,##0.00 €'
            row[4].number_format = '#,##0.00 €'

        # Total sum
        total_sum = sum(data['Total Price'] for data in invoice_data)
        ws.append(['', '', '', 'Total Price', total_sum])
        total_cell = ws.cell(row=ws.max_row, column=5)
        total_cell.font = bold_font
        total_cell.number_format = '#,##0.00 €'

        wb.save(invoice_path)

    def process_incoming_goods(self, filename):
        """Processes incoming goods from a CSV file.

        Args:
            filename (str): Path to the CSV file containing incoming goods data.
        """
        conn, cursor = self.connect_db()
        with open(filename, newline='', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                product_data = row.get('CODECONTENT', None)
                product_quantity = row.get('QUANTITY', None)
                if product_data:
                    product_info = product_data.split(', ')
                    product_id = int(product_info[0])
                    cursor.execute('SELECT * FROM products WHERE id = ?', (product_id,))
                    existing_product = cursor.fetchone()

                    if existing_product:
                        new_quantity = existing_product[2] + int(product_quantity)
                        cursor.execute('UPDATE products SET quantity = ? WHERE id = ?', (new_quantity, product_id))
                    else:
                        cursor.execute('''
                        INSERT INTO products (id, product_name, quantity, price, description)
                        VALUES (?, ?, ?, ?, ?)
                        ''', (product_id, product_info[1], int(product_quantity), float(product_info[2]), product_info[3]))

        shutil.move(filename, "./IncomingGoods/Processed")
        conn.commit()
        conn.close()
